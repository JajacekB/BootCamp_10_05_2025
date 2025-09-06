import glob
from pprint import pprint
from openpyxl import load_workbook


file_list = glob.glob("*.xlsx")
print(file_list)
pprint(file_list)

for file in file_list:
    try:
        wb = load_workbook(file)
        ws = wb.active

        value = ws['A1'].value
        print(f"{file}: A1 = {value}")
    except Exception as e:
        print("błąd:", e)
