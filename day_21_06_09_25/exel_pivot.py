from openpyxl import Workbook

# Przygotowanie danych
data = [
    ["Data", "Kategoria", "Produkt", "Sprzedaj"],
    ["2024-01-01", "Elektronika", "Telefon", 1500],
    ["2024-01-02", "Elektronika", "Telefon", 1500],
    ["2024-01-02", "Odzież", "Koszula", 50],
    ["2024-01-03", "Odzież", "Kurtka", 200],
    ["2024-01-04", "Elektronika", "Laptop", 1500],
]

# Tworzymy plik Excel
wb = Workbook()
ws = wb.active
ws.title = "Dane"

# Zapisujemy dane do arkusza "Dane"
for row in data:
    ws.append(row)

# Przygotowanie do stworzenia tabeli przestawnej
# Grupowanie danych po "Kategoria" i "Produkt"
pivot_data = {}
for row in data[1:]:
    category = row[1]
    product = row[2]
    sales = row[3]

    # Tworzymy zagnieżdżony słownik: {Kategoria: {Produkt: suma sprzedanych}}
    if category not in pivot_data:
        pivot_data[category] = {}
    if product not in pivot_data[category]:
        pivot_data[category][product] = 0
    pivot_data[category][product] += sales

# Tworzymy arkusz "Tabela Przestawna"
pivot_ws = wb.create_sheet("Tabela Przestawna")

# Nagłówki tabeli przestawnej
pivot_ws.cell(row=1, column=1, value="Kategoria")
products = set(product for category in pivot_data.values() for product in category.keys())
products = sorted(products)  # Posortowanie produktów
for col_idx, product in enumerate(products, 2):
    pivot_ws.cell(row=1, column=col_idx, value=product)

# Zapisujemy dane tabeli przestawnej
row_idx = 2  # Startujemy od drugiego wiersza
for category, products_data in pivot_data.items():
    pivot_ws.cell(row=row_idx, column=1, value=category)  # Kategoria w pierwszej kolumnie
    for col_idx, product in enumerate(products, 2):
        # Wartość sprzedaży dla danego produktu
        sales = products_data.get(product, 0)
        pivot_ws.cell(row=row_idx, column=col_idx, value=sales)
    row_idx += 1

# Zapisz plik
wb.save("tabela_przestawna.xlsx")
